import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table,
    TableStyle
)
from reportlab.lib.styles import getSampleStyleSheet

class SalesDataProcessor:
    """
    Class to process sales data.

    This class calculates the revenue for each product,
    sorts the products by revenue (descending order),
    and computes the total revenue.
    """
    def __init__(self, sales_data: dict):
        """
        Initialize with the raw sales data.

        Args:
            sales_data (dict): Dictionary with product names as keys and a dict of 
                               quantity and price as values.
        """
        self.sales_data = sales_data
        self.sorted_data = []
        self.total_revenue = 0

    def process_data(self):
        """
        Process the sales data to calculate revenue and sort the data.

        Returns:
            tuple: A tuple containing:
                - sorted_data (list of dict): List of dictionaries with keys
                  'product', 'quantity', 'price', and 'revenue', sorted in descending order by revenue.
                - total_revenue (float): Total revenue from all products.
        """
        data_list = []
        total = 0
        for product, info in self.sales_data.items():
            try:
                quantity = info.get("quantity", 0)
                price = info.get("price", 0)
                revenue = quantity * price
            except Exception as e:
                print(f"Error processing product {product}: {e}")
                continue

            data_list.append({
                "product": product,
                "quantity": quantity,
                "price": price,
                "revenue": revenue
            })
            total += revenue

        # Sort the list of dictionaries by revenue in descending order.
        self.sorted_data = sorted(data_list, key=lambda x: x["revenue"], reverse=True)
        self.total_revenue = total

        return self.sorted_data, self.total_revenue


class PDFReportGenerator:
    """
    Class to generate a PDF report from processed sales data.

    The report includes a company logo, a title, a table of sales data,
    and a summary row for the total revenue.
    """
    def __init__(self, output_filename: str, logo_path: str = None, title: str = "Sales Report 2024"):
        """
        Initialize the PDF generator with an output filename, optional logo path, and title.

        Args:
            output_filename (str): The filename for the generated PDF.
            logo_path (str, optional): Path to the company logo image file.
            title (str): Title of the sales report.
        """
        self.output_filename = output_filename
        self.logo_path = logo_path
        self.title = title
        self.styles = getSampleStyleSheet()

    def generate_report(self, sorted_data, total_revenue):
        """
        Generate the PDF report using the processed sales data.

        Args:
            sorted_data (list of dict): List of processed sales data.
            total_revenue (float): Total revenue to display in the summary.
        """
        # Create the PDF document.
        doc = SimpleDocTemplate(self.output_filename, pagesize=letter)
        elements = []

        # Create the header with logo and title in one row.
        header_data = []
        # Prepare the logo cell.
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                logo = Image(self.logo_path, width=100, height=50)
            except Exception as e:
                print(f"Error loading logo: {e}")
                logo = Paragraph("", self.styles["Normal"])
        else:
            logo = Paragraph("", self.styles["Normal"])
        
        # Prepare the title cell.
        title_paragraph = Paragraph(self.title, self.styles["Title"])
        
        # Combine them into a row.
        header_data.append([logo, title_paragraph])
        
        # Create a table for the header.
        header_table = Table(header_data, colWidths=[110, 400])
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, 0), "CENTER")
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 24))

        # Prepare table data with header.
        table_data = []
        table_data.append(["Product", "Quantity", "Price ($)", "Revenue ($)"])

        # Add a row for each product.
        for item in sorted_data:
            table_data.append([
                item["product"],
                item["quantity"],
                f"{item['price']:.2f}",
                f"{item['revenue']:.2f}"
            ])

        # Append the summary row with total revenue.
        table_data.append(["Total Revenue:", "", "", f"{total_revenue:.2f}"])

        # Create the table.
        table = Table(table_data, hAlign="LEFT")
        
        # Define table styling.
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.gray),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -2), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black)
        ])
        # Make the total revenue row bold.
        style.add("FONTNAME", (0, len(table_data)-1), (-1, len(table_data)-1), "Helvetica-Bold")
        table.setStyle(style)

        # Append the table to the elements.
        elements.append(table)

        # Build the PDF document.
        try:
            doc.build(elements)
            print(f"PDF report generated successfully: {self.output_filename}")
        except Exception as e:
            print(f"Error generating PDF: {e}")


class ExcelReportGenerator:
    """
    Class to export sales data to an Excel file.

    The Excel file will include a title, a table of sorted sales data,
    and a summary row for the total revenue.
    """
    def __init__(self, output_filename: str, title: str = "Sales Report 2024"):
        """
        Initialize the Excel report generator with an output filename and title.

        Args:
            output_filename (str): The filename for the generated Excel file.
            title (str): Title of the sales report.
        """
        self.output_filename = output_filename
        self.title = title

    def generate_report(self, sorted_data, total_revenue):
        """
        Generate the Excel report using the processed sales data.

        Args:
            sorted_data (list of dict): List of processed sales data.
            total_revenue (float): Total revenue to display in the summary.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            print("Please install openpyxl to export data to Excel: pip install openpyxl")
            return

        # Create a new workbook and select the active worksheet.
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Data"

        # Write the report title in the first row and merge cells A1:D1.
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = self.title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Leave a blank row (row 2) and add table headers in row 3.
        headers = ["Product", "Quantity", "Price ($)", "Revenue ($)"]
        ws.append([])  # Row 2 (empty)
        ws.append(headers)  # Row 3

        # Write each product's data to the worksheet.
        for item in sorted_data:
            ws.append([
                item["product"],
                item["quantity"],
                f"{item['price']:.2f}",
                f"{item['revenue']:.2f}"
            ])

        # Append the summary row with total revenue.
        ws.append(["Total Revenue:", "", "", f"{total_revenue:.2f}"])

        # Optionally, style the header row (row 3) to be bold.
        header_font = Font(bold=True)
        for cell in ws[3]:
            cell.font = header_font

        # Save the workbook to the specified file.
        try:
            wb.save(self.output_filename)
            print(f"Excel report generated successfully: {self.output_filename}")
        except Exception as e:
            print(f"Error generating Excel report: {e}")


if __name__ == "__main__":
    # Define the sample sales data.
    sales_data = {
        "Mouse": {"quantity": 10, "price": 20},
        "Laptop": {"quantity": 5, "price": 800},
        "Keyboard": {"quantity": 7, "price": 50},
        "Monitor": {"quantity": 3, "price": 300},
        "USB Drive": {"quantity": 15, "price": 10}
    }

    # Process the sales data.
    processor = SalesDataProcessor(sales_data)
    sorted_data, total_revenue = processor.process_data()

    # Define the output filenames and logo image path.
    output_pdf = "sales_report.pdf"
    output_excel = "sales_report.xlsx"
    logo_image = "company_logo.png"  # Update the path as needed or set to None if not available.

    # Generate the PDF report.
    pdf_generator = PDFReportGenerator(output_pdf, logo_image, "Sales Report 2024")
    pdf_generator.generate_report(sorted_data, total_revenue)

    # Generate the Excel report.
    excel_generator = ExcelReportGenerator(output_excel, "Sales Report 2024")
    excel_generator.generate_report(sorted_data, total_revenue)
