class ExcelReportGenerator:
    def __init__(self, output_filename: str, title: str = "Sales Report 2024"):
        self.output_filename = output_filename
        self.title = title

    def generate_report(self, sorted_data, total_revenue):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            print("Please install openpyxl to export data to Excel: pip install openpyxl")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Data"

        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = self.title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        headers = ["Product", "Quantity", "Price ($)", "Revenue ($)"]
        ws.append([])  # Row 2 (empty)
        ws.append(headers)  # Row 3

        # Modified: Access ProductSale object attributes directly
        for item in sorted_data:
            ws.append([
                item.product,
                item.quantity,
                f"{item.price:.2f}",
                f"{item.revenue:.2f}"
            ])

        ws.append(["Total Revenue:", "", "", f"{total_revenue:.2f}"])

        header_font = Font(bold=True)
        for cell in ws[3]:
            cell.font = header_font

        try:
            wb.save(self.output_filename)
            print(f"Excel report generated successfully: {self.output_filename}")
        except Exception as e:
            print(f"Error generating Excel report: {e}")